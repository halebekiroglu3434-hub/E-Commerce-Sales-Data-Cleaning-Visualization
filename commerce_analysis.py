import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


print(os.getcwd())

data = pd.read_csv(
    "ecommerce_log.csv",
    keep_default_na=False
)

df = pd.DataFrame(data)

print(df.head(9))
print(df.shape)
print(df.isnull().sum())

logging.basicConfig(level=logging.INFO)
#--date normalization--

#df.copy

df["OrderDate_Parsed"] = df["OrderDate"]


def pre_clean(date_str):
    df["OrderDate_Parsed"] = df["OrderDate_Parsed"].replace("-", "/").replace(".", "/")
    date_str = str(date_str).strip().lower()
    if date_str in ['unknown', 'not available', 'nan', 'none', '']:
        return None
    return date_str

dates = [pre_clean(date) for date in df["OrderDate_Parsed"]]
df["OrderDate_Parsed"] = pd.to_datetime(
    dates,
    dayfirst=False,
    errors='coerce',
    format='mixed'
)

print(df["OrderDate_Parsed"])

print(df.info)

print(df["Category"])


df["Category_Norm"] = (
    df["Category"]
    .str.strip()
    .str.lower()
)

category_map = {
    "electronics": "Electronics",
    "eletronics": "Electronics",
    "elektronics": "Electronics",
    "elec.": "Electronics",
    "elex.": "Electronics",

    "office supplies": "Office Supplies",
    "officesupply": "Office Supplies",

    "home&kitchen": "Home & Kitchen",
    "Home & Kitchen": "Home & Kitchen",
}

df["Category_Parsed"] = df["Category_Norm"].map(category_map)


#Alternative : df["Category_Parsed"] = df["Category"].replace("elec.", "Electronics").replace("eletronics", "Electronics").replace("elektronics", "Electronics").replace("Elex.","Electronics").replace('Office supplies', "Officesupplies").replace("Home&Kitchen","Home & Kitchen").str.strip().str.capitalize()

print(df["Category_Parsed"].unique())
print(df["Category"].unique())

print(df["UnitPrice"])

def clean_currency(value):
    if value is None:
        return None

    v = str(value).strip()


    if v.lower() == "free" or v == "":
        return 0.0

    for token in ["$", "USD", "TL"]:
        v = v.replace(token, "")

    v = v.strip()

    # 🔑 Normalization
    last_comma_index = v.rfind(",")
    last_dot_index = v.rfind(".")

    #(1.200,50 or 1,200.50)
    if last_comma_index != -1 and last_dot_index != -1:
        if last_comma_index > last_dot_index:

            v = v.replace(".", "").replace(",", ".")
        else:
            #(1,200.50) -> US
            v = v.replace(",", "")

    #(12,50 or 1,200)
    elif last_comma_index != -1:
        # (universal)
        if len(v) - last_comma_index - 1 == 2:
            v = v.replace(",", ".")  # 12,50 -> 12.50
        else:
            v = v.replace(",", "")  # 1,200 -> 1200

    #(Standart float) -> Same

    try:
        return float(v)
    except ValueError:
        return None

df["UnitPrice_numeric"] = df["UnitPrice"].apply(clean_currency)

print(df["UnitPrice_numeric"].value_counts())


def detect_currency(value):
    v = str(value).strip().lower()

    if "$" in v or "usd" in v:
        return v.replace(v, "USD")

    elif "tl" in v:
        return "TL"

    else:
        return "Unknown"

df["UnitPrice_Currency"] = df["UnitPrice"].apply(detect_currency)

#print(df["UnitPrice"].value_counts())
#print(df["UnitPrice_Currency"].value_counts())

Quantity = df["Quantity"].apply(clean_currency)
TotalPrice = df["TotalPrice"].apply(clean_currency)

df["TotalPrice_Final"] = df["UnitPrice_numeric"] * Quantity#.astype(float)
mistakes = TotalPrice - df["TotalPrice_Final"]
print(mistakes)
logging.warning("Total prices have been successfully Checked")
#print(df.columns)


print(df["TotalPrice_Final"])
print(df["UnitPrice_numeric"])

#----Pre Visualization Adjustments----#

vis_df = df[["UnitPrice_numeric", "TotalPrice_Final", "UnitPrice_Currency", "Quantity"]]


EXCHANGE_RATE_TL_TO_USD = 1 / 30

#logging.INFO("We could have use real time exchange")
def normalize_price(row):
    currceny = str(row["UnitPrice_Currency"]).upper()
    amount = row["TotalPrice_Final"]

    if "USD" in currceny:
        return amount
    elif "TL" in currceny:
        return amount * EXCHANGE_RATE_TL_TO_USD
    else:
        return None

vis_df = df[
    ["Category", "UnitPrice_numeric", "TotalPrice_Final", "UnitPrice_Currency", "Quantity"]
]

df["Total_Revenue_USD"] = df.apply(normalize_price, axis=1)


Status = (
    df.groupby(["Category_Parsed", "OrderStatus"])
      .size()
      .unstack(fill_value=0)

)
#----Visualization----#

Status.plot(kind="bar", stacked=True, figsize=(10, 5))
plt.title("Order Status Distribution by Category")
plt.xlabel("Category_Parsed")
plt.ylabel("Number of Orders")
plt.xticks(rotation=0)
plt.tight_layout()

plt.savefig(
    "order_status_distribution.png",
    dpi=300,
    bbox_inches="tight"
)
plt.show()
#------


plt.figure(figsize=(18, 6))

# (Pie Chart)
plt.subplot(1, 3, 1)
currency_counts = df['UnitPrice_Currency'].value_counts()
plt.pie(currency_counts, labels=currency_counts.index, autopct='%1.1f%%', startangle=140, colors=['#ff9999','#66b3ff','#99ff99'])
plt.title('Currency Distribution')

#(Barplot)
plt.subplot(1, 3, 2)
sns.barplot(x='Category_Parsed', y='Total_Revenue_USD', data=df, estimator=sum, errorbar=None, palette='Blues_d', hue='Category_Parsed')
plt.title('Category-Based Revenue($)')
plt.xticks(rotation=45)
plt.ylabel('Total Income(USDD)')

# Price vs Quantity (Scatterplot)
plt.subplot(1, 3, 3)

sns.scatterplot(x='UnitPrice_numeric', y='Quantity', hue='Category_Parsed', data=df, s=100, alpha=0.7)
plt.title('Fare vs Quantity')
plt.xlabel('Unit Price')
plt.ylabel('Quantity')
plt.xscale('log')

plt.tight_layout()

output_path = os.path.join(
    os.getcwd(),
    "ecommerce_analysis_dashboard_v2.png"
)

plt.savefig(
    "ecommerce_analysis_dashboard_v.png",
    dpi=300,
    bbox_inches="tight"
)
plt.show()

logging.info("Saved to:", output_path)
#---Fınal Excell Format---#

file_name = "ecommerce_data_analysis.xlsx"
df.to_excel(file_name, index=False)

wb = load_workbook(file_name)
ws = wb.active

header_fill_blue = PatternFill(
    start_color='4F81BD',
    end_color='4F81BD',
    fill_type='solid',
)
header_fill_green = PatternFill(
    start_color="04e06f",
    end_color="04e06f",
    fill_type="solid"
)

header_font = Font(
    color="FFFFFF",
    bold=True,
)
for i, cell in enumerate(ws[1]):
    if i % 2 == 0:
        cell.fill = header_fill_green
    else:
        cell.fill = header_fill_blue

    cell.font = header_font
wb.save(file_name)

logging.info("File saved successfully")
